# -*- coding: utf-8 -*-
import io
import datetime as dt
from typing import Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

WON_FORMAT = '_-₩* #,##0_-;_-₩* -#,##0_-;_-₩* "-"_-;_-@_-'

def _apply_table_style(ws, header_row: int, ncols: int, freeze_row: int):
    # 헤더 스타일
    header_fill = PatternFill("solid", fgColor="1F4E79")  # 진한 파랑
    header_font = Font(color="FFFFFF", bold=True)
    align = Alignment(vertical="center", horizontal="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
        cell.border = border

    ws.freeze_panes = ws[f"A{freeze_row}"]

def _autosize(ws, max_col: int, min_width=10, max_width=28):
    for c in range(1, max_col + 1):
        col = get_column_letter(c)
        max_len = 0
        for cell in ws[col]:
            if cell.value is None:
                continue
            v = str(cell.value)
            max_len = max(max_len, len(v))
        width = max(min_width, min(max_width, max_len + 2))
        ws.column_dimensions[col].width = width

def _write_df(ws, df: pd.DataFrame, title: str, start_row: int = 1, money_col: Optional[str] = None):
    # 타이틀
    ws.cell(row=start_row, column=1, value=title).font = Font(size=16, bold=True)
    ws.cell(row=start_row, column=1).alignment = Alignment(vertical="center", horizontal="left")
    start_row += 2

    # 헤더
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    header_row = start_row
    start_row += 1

    # 데이터
    for i, (_, r) in enumerate(df.iterrows(), start=start_row):
        for j, col in enumerate(df.columns, start=1):
            v = r[col]
            if isinstance(v, dt.date):
                ws.cell(row=i, column=j, value=v)
                ws.cell(row=i, column=j).number_format = "yyyy-mm-dd"
            else:
                ws.cell(row=i, column=j, value=None if (pd.isna(v) if hasattr(pd, "isna") else False) else v)

    # 금액 서식
    if money_col and money_col in df.columns and not df.empty:
        m_idx = list(df.columns).index(money_col) + 1
        for i in range(header_row + 1, header_row + 1 + len(df)):
            ws.cell(row=i, column=m_idx).number_format = WON_FORMAT
            ws.cell(row=i, column=m_idx).alignment = Alignment(horizontal="right")

    # 테이블 스타일
    _apply_table_style(ws, header_row=header_row, ncols=len(df.columns), freeze_row=header_row + 1)

    # 데이터 셀 기본 정렬 + 약한 테두리(가독성)
    thin = Side(style="thin", color="E6E6E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + len(df), min_col=1, max_col=len(df.columns)):
        for cell in row:
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            cell.border = border

    _autosize(ws, len(df.columns))

def export_day_xlsx(
    d: dt.date,
    income_df: pd.DataFrame,
    expense_df: pd.DataFrame,
    church_name: str = "평안한교회",
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("수입")
    ws2 = wb.create_sheet("지출")

    _write_df(ws1, income_df, f"{d.isoformat()} 수입 장부 ({church_name})", money_col="금액")
    _write_df(ws2, expense_df, f"{d.isoformat()} 지출 장부 ({church_name})", money_col="금액")

    # 요약 시트
    ws3 = wb.create_sheet("요약")
    ws3.cell(row=1, column=1, value=f"{d.isoformat()} 재정 요약 ({church_name})").font = Font(size=16, bold=True)

    income_total = float(pd.to_numeric(income_df["금액"], errors="coerce").fillna(0).sum()) if not income_df.empty and "금액" in income_df.columns else 0.0
    expense_total = float(pd.to_numeric(expense_df["금액"], errors="coerce").fillna(0).sum()) if not expense_df.empty and "금액" in expense_df.columns else 0.0

    ws3.cell(row=3, column=1, value="수입 합계").font = Font(bold=True)
    ws3.cell(row=3, column=2, value=income_total).number_format = WON_FORMAT
    ws3.cell(row=4, column=1, value="지출 합계").font = Font(bold=True)
    ws3.cell(row=4, column=2, value=expense_total).number_format = WON_FORMAT
    ws3.cell(row=5, column=1, value="차액(수입-지출)").font = Font(bold=True)
    ws3.cell(row=5, column=2, value=income_total - expense_total).number_format = WON_FORMAT

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 18
    for r in range(3, 6):
        ws3.cell(row=r, column=2).alignment = Alignment(horizontal="right")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

def export_all_xlsx(income_all: pd.DataFrame, expense_all: pd.DataFrame, church_name: str = "평안한교회") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("수입전체")
    ws2 = wb.create_sheet("지출전체")

    _write_df(ws1, income_all, f"수입 전체 데이터 ({church_name})", money_col="금액")
    _write_df(ws2, expense_all, f"지출 전체 데이터 ({church_name})", money_col="금액")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_tables_xlsx(filename_prefix: str, sheets: dict, money_columns: list[str] | None = None) -> bytes:
    """
    여러 표(DataFrame)를 한 번에 엑셀로 내보냅니다.
    sheets: {sheet_name: dataframe}
    money_columns: 금액 서식을 적용할 컬럼명 리스트(해당 컬럼이 존재할 때만)
    """
    import io
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment

    money_columns = money_columns or []
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        if df is None:
            df = pd.DataFrame()

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            ws.append(row)
            if r_idx == 1:
                for c_idx in range(1, len(row) + 1):
                    cell = ws.cell(row=1, column=c_idx)
                    cell.font = header_font
                    cell.alignment = center

        # column widths
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            for cell in col[: min(len(col), 200)]:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 28)

        # money format
        if not df.empty:
            cols = list(df.columns)
            for col_name in money_columns:
                if col_name in cols:
                    idx = cols.index(col_name) + 1
                    for r in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r, column=idx)
                        if isinstance(cell.value, (int, float)) and cell.value is not None:
                            cell.number_format = WON_FORMAT
                            cell.alignment = right

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
